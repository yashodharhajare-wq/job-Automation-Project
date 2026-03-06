import os
import shutil
from datetime import datetime
from docx import Document
from docx.shared import RGBColor
from openpyxl import Workbook, load_workbook
import webbrowser

# ---- USER INPUT ----
company = input("Enter Company Name: ")
job_title = input("Enter Job Title: ")
Location = input("Enter Location: ")

# ---- CREATE NUMBERED FOLDER ----
existing_numbers = []

for name in os.listdir():
    if os.path.isdir(name):
        parts = name.split(" ", 1)  # Look for "NN something"
        if parts and parts[0].isdigit():
            existing_numbers.append(int(parts[0]))

next_number = max(existing_numbers, default=0) + 1

# Only sanitize the right part (company + job title), keep the "NN " intact
right_part = f"{company}_{job_title}".replace(" ", "_").replace("/", "")
folder_name = f"{next_number:02d} {right_part}"
os.makedirs(folder_name, exist_ok=True)
folder = folder_name


# ---- COPY & FORMAT COVER LETTER ----
template_path = "templates/Yashodhar Hajare_ Anschreiben.docx"
destination_path = os.path.join(folder, "Yashodhar Hajare_ Anschreiben.docx")

doc = Document(template_path)

for paragraph in doc.paragraphs:
    full_text = paragraph.text

    if "[COMPANY_NAME]" in full_text:
        paragraph.text = full_text.replace("[COMPANY_NAME]", company)
        for run in paragraph.runs:
            run.font.name = "Calibri"
            run.font.color.rgb = RGBColor(0x2E, 0x74, 0xB5)
            run.bold = True

    if "[JOB_TITLE]" in full_text:
        paragraph.text = paragraph.text.replace("[JOB_TITLE]", job_title)
        for run in paragraph.runs:
            run.font.name = "Calibri"
            run.font.color.rgb = RGBColor(0x2E, 0x74, 0xB5)
            run.bold = True

    if "[Location]" in full_text:
        paragraph.text = paragraph.text.replace("[Location]", Location)
        for run in paragraph.runs:
            run.font.name = "Calibri"
            run.font.color.rgb = RGBColor(0x2E, 0x74, 0xB5)
            run.bold = True

doc.save(destination_path)

# ---- OPEN CHROME SEARCH ----
search_query = f"{company} {job_title}"
webbrowser.open(f"https://www.google.com/search?q={search_query}")

# ---- ASK FOR JOB LINK ----
job_link = input("\nPaste Job Link Here: ")

# ---- UPDATE EXCEL ----
tracker_file = "Application_Tracker.xlsx"

if not os.path.exists(tracker_file):
    wb = Workbook()
    ws = wb.active
    ws.append(["SR No", "Date", "Company Name", "Job Title", "Location", "Job Link"])
    wb.save(tracker_file)

wb = load_workbook(tracker_file)
ws = wb.active

# Get correct SR No even after manual deletions
sr_numbers = [cell.value for cell in ws["A"][1:] if isinstance(cell.value, int)]
sr_no = max(sr_numbers, default=0) + 1

new_row = ws.max_row + 1

ws.cell(row=new_row, column=1, value=sr_no)
ws.cell(row=new_row, column=2, value=datetime.now().strftime("%Y-%m-%d"))
ws.cell(row=new_row, column=3, value=company)
ws.cell(row=new_row, column=4, value=job_title)
ws.cell(row=new_row, column=5, value=Location)

link_cell = ws.cell(row=new_row, column=6, value="Open Job")
link_cell.hyperlink = job_link
link_cell.style = "Hyperlink"

wb.save(tracker_file)

print("\n✔ Folder created")
print("✔ Cover letter formatted with colors and font")
print("✔ Chrome search opened")
print("✔ Excel tracker updated with job link")
