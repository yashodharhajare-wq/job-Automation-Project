from pathlib import Path
import shutil
from datetime import datetime
from docx import Document
from docx.shared import RGBColor
from openpyxl import Workbook, load_workbook
import webbrowser

# ---- PATH SETUP ----
BASE_DIR = Path(__file__).resolve().parent.parent

DATA_DIR = BASE_DIR / "data"
TEMPLATES_DIR = BASE_DIR / "templates"

EXCEL_FILE = DATA_DIR / "application_tracker.xlsx"
TEMPLATE_FILE = TEMPLATES_DIR / "cover_letter_template.docx"

DATA_DIR.mkdir(exist_ok=True)

# ---- USER INPUT ----
company = input("Enter Company Name: ")
job_title = input("Enter Job Title: ")
location = input("Enter Location: ")

# ---- CREATE NUMBERED FOLDER INSIDE DATA ----
existing_numbers = []

for folder in DATA_DIR.iterdir():
    if folder.is_dir():
        parts = folder.name.split(" ", 1)
        if parts and parts[0].isdigit():
            existing_numbers.append(int(parts[0]))

next_number = max(existing_numbers, default=0) + 1

right_part = f"{company}_{job_title}".replace(" ", "_").replace("/", "")
folder_name = f"{next_number:02d} {right_part}"

job_folder = DATA_DIR / folder_name
job_folder.mkdir(parents=True, exist_ok=True)

# ---- COPY & FORMAT COVER LETTER ----
destination_path = job_folder / "Anschreiben_yashodhar_Hajare.docx"

doc = Document(TEMPLATE_FILE)

for paragraph in doc.paragraphs:
    full_text = paragraph.text

    if "[COMPANY_NAME]" in full_text:
        paragraph.text = full_text.replace("[COMPANY_NAME]", company)

    if "[JOB_TITLE]" in paragraph.text:
        paragraph.text = paragraph.text.replace("[JOB_TITLE]", job_title)

    if "[LOCATION]" in paragraph.text:
        paragraph.text = paragraph.text.replace("[LOCATION]", location)

    # Apply formatting after replacement
    for run in paragraph.runs:
        if any(keyword in run.text for keyword in [company, job_title, location]):
            run.font.name = "Calibri"
            run.font.color.rgb = RGBColor(0x2E, 0x74, 0xB5)
            run.bold = True

doc.save(destination_path)


# ---- ASK FOR JOB LINK ----
job_link = input("\nPaste Job Link Here: ")

# ---- UPDATE EXCEL ----

tracker_file = "Application_Tracker.xlsx"

if not EXCEL_FILE.exists():
    wb = Workbook()
    ws = wb.active
    ws.append(["SR No", "Date", "Company Name", "Job Title", "Location", "Job Link"])
    wb.save(EXCEL_FILE)

wb = load_workbook(EXCEL_FILE)
ws = wb.active

sr_numbers = [cell.value for cell in ws["A"][1:] if isinstance(cell.value, int)]
sr_no = max(sr_numbers, default=0) + 1

new_row = ws.max_row + 1

ws.cell(row=new_row, column=1, value=sr_no)
ws.cell(row=new_row, column=2, value=datetime.now().strftime("%Y-%m-%d"))
ws.cell(row=new_row, column=3, value=company)
ws.cell(row=new_row, column=4, value=job_title)
ws.cell(row=new_row, column=5, value=location)

link_cell = ws.cell(row=new_row, column=6, value="Open Job")
link_cell.hyperlink = job_link
link_cell.style = "Hyperlink"

wb.save(EXCEL_FILE)

print("\n✔ Folder created")
print("✔ Cover letter generated")
print("✔ Chrome search opened")
print("✔ Excel tracker updated")
